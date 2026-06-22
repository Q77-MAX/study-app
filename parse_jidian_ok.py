import openpyxl
import json
import re

wb = openpyxl.load_workbook(r'C:\Users\74476\OneDrive\桌面\继电保护题库_OK.xlsx')
ws = wb[wb.sheetnames[0]]

TYPE_MAP = {4: 'single', 5: 'essay', 6: 'judge', 7: 'multiple'}
DIFF_MAP = {'易': 1, '中': 3, '难': 5, '': 1}

def parse_options(raw):
    if not raw or not str(raw).strip():
        return []
    parts = [p.strip() for p in str(raw).split('|') if p.strip()]
    result = []
    for p in parts:
        m = re.match(r'([A-H])[\.\、）\)\s]+(.+)', p)
        if m:
            result.append({'label': m.group(1), 'text': m.group(2).strip()})
        else:
            result.append({'label': None, 'text': p.strip()})
    return result

questions = []
stats = {}

for r in range(2, ws.max_row + 1):
    content = str(ws.cell(r, 2).value or '').strip()
    if not content:
        continue

    # Determine question type from which answer column is filled
    qtype = None
    answer = ''
    for c in [4, 5, 6, 7]:
        v = ws.cell(r, c).value
        if v is not None and str(v).strip():
            qtype = TYPE_MAP[c]
            answer = str(v).strip()
            break

    if qtype is None:
        continue

    # Parse options from Col3
    opts_raw = str(ws.cell(r, 3).value or '')
    opts_with_labels = parse_options(opts_raw)
    options = [o['text'] for o in opts_with_labels]

    # Normalize answer
    if qtype == 'single':
        # Handle multi-line garbage: take first line if it's a single letter
        first_line = answer.split('\n')[0].strip().upper()
        if len(first_line) == 1 and first_line.isalpha():
            answer = first_line
        else:
            answer = re.sub(r'[^A-H]', '', answer.upper())
    elif qtype == 'multiple':
        answer = re.sub(r'[^A-H]', '', answer.upper())
        answer = ''.join(sorted(answer))
    elif qtype == 'judge':
        if answer in ['对', '正确', '是', '√', '✓']:
            answer = 'A'
        elif answer in ['错', '错误', '否', '×', '✗']:
            answer = 'B'
        options = ['对', '错']

    questions.append({
        'type': qtype,
        'content': content,
        'options': options,
        'answer': answer,
        'explanation': '',
        'knowledgePoints': [],
        'difficulty': 1,
    })
    stats[qtype] = stats.get(qtype, 0) + 1

# Validate
issues = []
for q in questions:
    if q['type'] == 'single' and q['answer'] and q['options']:
        idx = ord(q['answer']) - ord('A')
        if idx < 0 or idx >= len(q['options']):
            issues.append(f"Single: answer={q['answer']}, opts={len(q['options'])}: {q['content'][:60]}")
    if q['type'] == 'multiple' and q['answer'] and q['options']:
        for letter in q['answer']:
            idx = ord(letter) - ord('A')
            if idx < 0 or idx >= len(q['options']):
                issues.append(f"Multi: answer={q['answer']}, opts={len(q['options'])}: {q['content'][:60]}")

labels = {'single': '单选题', 'multiple': '多选题', 'judge': '判断题', 'fill': '填空题', 'essay': '简答题'}
print(f'Total: {len(questions)}')
for t, c in sorted(stats.items()):
    print(f'  {labels.get(t, t)} ({t}): {c}')
print(f'Issues: {len(issues)}')
for iss in issues[:5]:
    print(f'  {iss}')

# Sample
for qt in ['single', 'multiple', 'judge', 'essay']:
    for q in questions:
        if q['type'] == qt:
            print(f'\n[{qt}] Q: {q["content"][:120]}')
            if q['options']:
                for i, o in enumerate(q['options'][:6]):
                    print(f'  {chr(65+i)}. {o[:80]}')
            print(f'  Answer: {q["answer"][:100]}')
            break

output_path = r'C:\Users\74476\study-app\public\questions_jidianbaohu.json'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(questions, f, ensure_ascii=False, indent=2)
print(f'\nSaved to questions_jidianbaohu.json')
