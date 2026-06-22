import openpyxl
import json
import re

wb = openpyxl.load_workbook(r'C:\Users\74476\OneDrive\桌面\继电保护题库.xlsx')
ws = wb[wb.sheetnames[0]]

type_map = {
    '单选题': 'single',
    '多选题': 'multiple',
    '判断题': 'judge',
    '填空题': 'fill',
    '简答题': 'essay',
}
diff_map = {'易': 1, '中': 3, '难': 5}

def parse_options_with_labels(raw):
    if not raw or not raw.strip():
        return []
    parts = [p.strip() for p in raw.split('|') if p.strip()]
    result = []
    for p in parts:
        m = re.match(r'([A-H])[\.\、）\)\s]+(.+)', p)
        if m:
            result.append({'label': m.group(1), 'text': m.group(2).strip()})
        else:
            result.append({'label': None, 'text': p.strip()})
    return result

def normalize_options_and_answer(opts_with_labels, answer_letters, qtype):
    sorted_opts = sorted(opts_with_labels, key=lambda x: x['label'] or 'Z')
    options_texts = []
    old_to_new = {}
    for i, opt in enumerate(sorted_opts):
        options_texts.append(opt['text'])
        if opt['label']:
            old_to_new[opt['label']] = i

    new_answer_parts = []
    missing_labels = set()
    for letter in answer_letters:
        if letter in old_to_new:
            new_idx = old_to_new[letter]
            new_letter = chr(65 + new_idx)
            new_answer_parts.append(new_letter)
        else:
            missing_labels.add(letter)

    if qtype == 'single':
        new_answer = new_answer_parts[0] if new_answer_parts else answer_letters
    else:
        new_answer = ''.join(sorted(new_answer_parts))

    return options_texts, new_answer, missing_labels

questions = []
stats = {'single': 0, 'multiple': 0, 'judge': 0, 'fill': 0, 'essay': 0}
missing_label_issues = []

for r in range(2, ws.max_row + 1):
    raw_type = str(ws.cell(r, 2).value or '').strip()
    if not raw_type:
        continue

    qtype = type_map.get(raw_type)
    if not qtype:
        continue

    content = str(ws.cell(r, 3).value or '').strip()
    if not content:
        continue

    opts_raw = str(ws.cell(r, 9).value or '')
    opts_with_labels = parse_options_with_labels(opts_raw)

    raw_answer = ''
    if qtype == 'single':
        raw_answer = str(ws.cell(r, 4).value or '').strip().upper()
        raw_answer = re.sub(r'[^A-H]', '', raw_answer)
    elif qtype == 'multiple':
        raw_answer = str(ws.cell(r, 7).value or '').strip().upper()
        raw_answer = re.sub(r'[^A-H]', '', raw_answer)
    elif qtype == 'judge':
        raw_answer = str(ws.cell(r, 6).value or '').strip()
        if raw_answer in ['对', '正确', '是', '√', '✓']:
            raw_answer = 'A'
        elif raw_answer in ['错', '错误', '否', '×', '✗']:
            raw_answer = 'B'
    elif qtype == 'fill':
        raw_answer = str(ws.cell(r, 8).value or '').strip()
    elif qtype == 'essay':
        raw_answer = str(ws.cell(r, 5).value or '').strip()

    if qtype == 'judge':
        options = ['对', '错']
        answer = raw_answer
    elif qtype in ('single', 'multiple') and opts_with_labels:
        # Handle missing option A: insert placeholder before normalize
        existing_labels = {o['label'] for o in opts_with_labels if o['label']}
        needed_labels = set(raw_answer)
        missing_labels = needed_labels - existing_labels
        if missing_labels:
            # Insert placeholder for each missing label at the correct position
            for ml in sorted(missing_labels):
                insert_pos = ord(ml) - ord('A')
                # Clamp to valid range
                insert_pos = min(insert_pos, len(opts_with_labels))
                opts_with_labels.insert(insert_pos, {'label': ml, 'text': f'(选项{ml})'})

        options, answer, _ = normalize_options_and_answer(
            opts_with_labels, raw_answer, qtype
        )
    else:
        options = [o['text'] for o in opts_with_labels]
        answer = raw_answer

    diff_raw = str(ws.cell(r, 10).value or '').strip()
    difficulty = diff_map.get(diff_raw, 1)

    questions.append({
        'type': qtype,
        'content': content,
        'options': options,
        'answer': answer,
        'explanation': '',
        'knowledgePoints': [],
        'difficulty': difficulty,
    })
    stats[qtype] = stats.get(qtype, 0) + 1

# Write log
lines = []
lines.append(f'Total: {len(questions)} questions')
labels_cn = {'single': '单选题', 'multiple': '多选题', 'judge': '判断题', 'fill': '填空题', 'essay': '简答题'}
for t, c in stats.items():
    lines.append(f'  {labels_cn.get(t, t)} ({t}): {c}')

lines.append(f'\n=== Missing label issues: {len(missing_label_issues)} ===')
for iss in missing_label_issues[:15]:
    lines.append(f"\nRow {iss['row']} [{iss['type']}]: answer={iss['answer']}, missing={iss['missing']}")
    lines.append(f"  Q: {iss['content']}")
    lines.append(f"  Opts: {iss['opts_found']}")
    lines.append(f"  Normalized: answer={iss['norm_answer']}, opts={iss['norm_opts']}")

# Samples per type
lines.append('\n=== Samples ===')
for qt in ['single', 'multiple', 'judge', 'fill', 'essay']:
    for q in questions:
        if q['type'] == qt:
            lines.append(f'\n[{qt}] Q: {q["content"][:200]}')
            if q['options']:
                for i, o in enumerate(q['options'][:6]):
                    lines.append(f'  {chr(65+i)}. {o[:120]}')
            lines.append(f'  Answer: {q["answer"][:300]}')
            break

# Answer validation
answer_issues = []
for q in questions:
    if q['type'] == 'single' and q['answer'] and q['options']:
        idx = ord(q['answer']) - ord('A')
        if idx < 0 or idx >= len(q['options']):
            answer_issues.append(q)
    if q['type'] == 'multiple' and q['answer'] and q['options']:
        for letter in q['answer']:
            idx = ord(letter) - ord('A')
            if idx < 0 or idx >= len(q['options']):
                answer_issues.append(q)
                break

lines.append(f'\n=== Answer/index mismatch: {len(answer_issues)} ===')
for q in answer_issues[:10]:
    lines.append(f'  [{q["type"]}] Answer: {q["answer"]}, Opts: {len(q["options"])}')

with open(r'C:\Users\74476\study-app\jidian_convert_log.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

# Save
output_path = r'C:\Users\74476\study-app\public\questions_jidianbaohu.json'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(questions, f, ensure_ascii=False, indent=2)

print(f'Done. {len(questions)} saved. Stats: {stats}')
print(f'Missing labels: {len(missing_label_issues)}')
print(f'Answer mismatches: {len(answer_issues)}')
