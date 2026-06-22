import openpyxl
import json
import re

def parse_excel(filepath, output_path, type_cols):
    """
    type_cols: list of (col_index, type_name) tuples
    e.g. [(4,'single'), (5,'judge'), (6,'multiple')]
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]

    col_to_type = {c: t for c, t in type_cols}

    def parse_opts(raw):
        if not raw or not str(raw).strip():
            return []
        parts = [p.strip() for p in str(raw).split('|') if p.strip()]
        result = []
        for p in parts:
            m = re.match(r'([A-H])[\.\、）\)\s]+(.+)', p)
            if m:
                result.append({'label': m.group(1), 'text': m.group(2).strip()})
            else:
                result.append({'label': None, 'text': p.strip()})
        return result

    questions = []
    stats = {}

    for r in range(2, ws.max_row + 1):
        content = str(ws.cell(r, 2).value or '').strip()
        if not content:
            continue

        qtype = None
        answer = ''
        for c, t in type_cols:
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                qtype = t
                answer = str(v).strip()
                break

        if qtype is None:
            continue

        opts_raw = str(ws.cell(r, 3).value or '')
        opts_with_labels = parse_opts(opts_raw)
        options = [o['text'] for o in opts_with_labels]

        # Normalize
        if qtype == 'single':
            first_line = answer.split('\n')[0].strip().upper()
            if len(first_line) == 1 and first_line.isalpha():
                answer = first_line
            else:
                answer = re.sub(r'[^A-H]', '', answer.upper())
        elif qtype == 'multiple':
            answer = re.sub(r'[^A-H]', '', answer.upper())
            answer = ''.join(sorted(answer))
        elif qtype == 'judge':
            if answer in ['对', '正确', '是', '√', '✓']:
                answer = 'A'
            elif answer in ['错', '错误', '否', '×', '✗']:
                answer = 'B'
            options = ['对', '错']

        questions.append({
            'type': qtype,
            'content': content,
            'options': options,
            'answer': answer,
            'explanation': '',
            'knowledgePoints': [],
            'difficulty': 1,
        })
        stats[qtype] = stats.get(qtype, 0) + 1

    # Validate
    issues = 0
    for q in questions:
        if q['type'] in ('single', 'multiple') and q['answer'] and q['options']:
            for letter in q['answer']:
                idx = ord(letter) - ord('A')
                if idx < 0 or idx >= len(q['options']):
                    issues += 1
    print(f'  Issues: {issues}')

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(questions, f, ensure_ascii=False, indent=2)

    labels = {'single': '单选', 'multiple': '多选', 'judge': '判断', 'fill': '填空', 'essay': '简答'}
    for t, c in sorted(stats.items()):
        print(f'  {labels.get(t,t)} ({t}): {c}')
    print(f'  Total: {len(questions)}')
    return questions

# === 继电保护题库 (OK version) ===
# Cols: 4=single, 5=essay, 6=judge, 7=multiple
print('=== 继电保护题库 ===')
parse_excel(
    r'C:\Users\74476\OneDrive\桌面\继电保护题库_OK.xlsx',
    r'C:\Users\74476\study-app\public\questions_jidianbaohu.json',
    [(4, 'single'), (5, 'essay'), (6, 'judge'), (7, 'multiple')]
)

# === 发电厂变电站题库 ===
# Cols: 4=single, 5=judge, 6=multiple (no essay/fill column)
print('\n=== 发电厂变电站题库 ===')
parse_excel(
    r'C:\Users\74476\OneDrive\桌面\发电厂变电站题库.xlsx',
    r'C:\Users\74476\study-app\public\questions_fadianchang.json',
    [(4, 'single'), (5, 'judge'), (6, 'multiple')]
)
